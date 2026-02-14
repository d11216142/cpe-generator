from flask import Flask, render_template, request, jsonify, send_file
import requests
import re
import random
import csv
import io
import json
from datetime import datetime, timedelta
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from db_config import (
    save_multiple_cpe_to_database, 
    load_db_connections, 
    save_db_connections,
    test_db_connection,
    set_current_db_config,
    get_current_db_config
)

app = Flask(__name__)

# Common vendor names for random generation
COMMON_VENDORS = [
    'microsoft', 'google', 'apple', 'oracle', 'adobe', 'mozilla', 
    'cisco', 'ibm', 'intel', 'hp', 'dell', 'lenovo', 'asus',
    'samsung', 'sony', 'lg', 'vmware', 'redhat', 'canonical',
    'apache', 'nginx', 'nodejs', 'python', 'java', 'php'
]

PRODUCT_PREFIXES = ['server', 'client', 'pro', 'enterprise', 'professional', 'community', 'standard', 'ultimate']
PRODUCT_TYPES = ['suite', 'manager', 'viewer', 'editor', 'player', 'reader', 'browser', 'office', 'database', 'framework']

# CPE Dictionary - Sample CPE entries representing real-world software, hardware and OS
CPE_DICTIONARY = [
    # Applications (a)
    'cpe:2.3:a:microsoft:office:2019:*:*:*:*:*:*:*',
    'cpe:2.3:a:microsoft:office:2021:sp1:*:*:professional:*:*:*',
    'cpe:2.3:a:microsoft:edge:120.0.2210.91:*:*:*:*:*:*:*',
    'cpe:2.3:a:google:chrome:120.0.6099.129:*:*:*:*:*:*:*',
    'cpe:2.3:a:google:chrome:119.0.6045.199:stable:*:*:*:*:*:*',
    'cpe:2.3:a:mozilla:firefox:121.0:*:*:*:*:*:*:*',
    'cpe:2.3:a:mozilla:firefox:120.0.1:*:*:*:*:linux:*:*',
    'cpe:2.3:a:mozilla:thunderbird:115.6.0:*:*:*:*:*:*:*',
    'cpe:2.3:a:adobe:acrobat_reader:23.008.20458:*:*:*:*:*:*:*',
    'cpe:2.3:a:adobe:photoshop:24.7.1:*:*:*:*:*:*:*',
    'cpe:2.3:a:adobe:illustrator:28.0:*:*:*:*:*:*:*',
    'cpe:2.3:a:oracle:java:1.8.0:update391:*:*:*:*:*:*',
    'cpe:2.3:a:oracle:java:17.0.9:*:*:*:*:*:*:*',
    'cpe:2.3:a:oracle:mysql:8.0.35:*:*:*:*:*:*:*',
    'cpe:2.3:a:oracle:virtualbox:7.0.12:*:*:*:*:*:*:*',
    'cpe:2.3:a:apache:tomcat:10.1.17:*:*:*:*:*:*:*',
    'cpe:2.3:a:apache:http_server:2.4.58:*:*:*:*:*:*:*',
    'cpe:2.3:a:apache:maven:3.9.6:*:*:*:*:*:*:*',
    'cpe:2.3:a:python:python:3.12.1:*:*:*:*:*:*:*',
    'cpe:2.3:a:python:python:3.11.7:*:*:*:*:*:*:*',
    'cpe:2.3:a:nodejs:node.js:20.10.0:*:*:*:lts:*:*:*',
    'cpe:2.3:a:nodejs:node.js:21.5.0:*:*:*:*:*:*:*',
    'cpe:2.3:a:php:php:8.3.1:*:*:*:*:*:*:*',
    'cpe:2.3:a:php:php:8.2.14:*:*:*:*:*:*:*',
    'cpe:2.3:a:vmware:workstation:17.5.0:*:*:*:pro:*:*:*',
    'cpe:2.3:a:vmware:vsphere:8.0:update2:*:*:*:*:*:*',
    'cpe:2.3:a:cisco:webex:43.12.0.27982:*:*:*:*:*:*:*',
    'cpe:2.3:a:cisco:anyconnect:4.10.07061:*:*:*:*:*:*:*',
    'cpe:2.3:a:ibm:db2:11.5.8:*:*:*:*:linux:*:*',
    'cpe:2.3:a:ibm:websphere:9.0.5.15:*:*:*:*:*:*:*',
    'cpe:2.3:a:intel:graphics_driver:31.0.101.4972:*:*:*:*:windows:*:*',
    'cpe:2.3:a:nvidia:geforce_experience:3.27.0.112:*:*:*:*:*:*:*',
    'cpe:2.3:a:zoom:zoom:5.16.10.26186:*:*:*:*:windows:*:*',
    'cpe:2.3:a:slack:slack:4.36.140:*:*:*:*:*:*:*',
    'cpe:2.3:a:docker:docker:24.0.7:*:*:*:*:*:*:*',
    'cpe:2.3:a:git:git:2.43.0:*:*:*:*:*:*:*',
    'cpe:2.3:a:7-zip:7-zip:23.01:*:*:*:*:*:*:*',
    'cpe:2.3:a:videolan:vlc:3.0.20:*:*:*:*:*:*:*',
    'cpe:2.3:a:wireshark:wireshark:4.2.0:*:*:*:*:*:*:*',
    'cpe:2.3:a:notepad_plus_plus:notepad\\+\\+:8.6.2:*:*:*:*:*:*:*',
    'cpe:2.3:a:postman:postman:10.21.1:*:*:*:*:*:*:*',
    'cpe:2.3:a:jetbrains:intellij_idea:2023.3.2:*:*:*:community:*:*:*',
    'cpe:2.3:a:visual_studio_code:visual_studio_code:1.85.2:*:*:*:*:*:*:*',
    'cpe:2.3:a:spotify:spotify:1.2.26.1187:*:*:*:*:*:*:*',
    'cpe:2.3:a:steam:steam:1702689516:*:*:*:*:*:*:*',
    'cpe:2.3:a:discord:discord:0.0.309:*:*:*:*:*:*:*',
    'cpe:2.3:a:malwarebytes:anti-malware:4.6.3:*:*:*:premium:*:*:*',
    'cpe:2.3:a:ccleaner:ccleaner:6.19.10858:*:*:*:free:*:*:*',
    
    # Operating Systems (o)
    'cpe:2.3:o:microsoft:windows_10:21h2:*:*:*:*:*:*:*',
    'cpe:2.3:o:microsoft:windows_11:22h2:*:*:*:*:*:*:*',
    'cpe:2.3:o:microsoft:windows_server_2022:*:*:*:*:*:*:*:*',
    'cpe:2.3:o:microsoft:windows_server_2019:*:*:*:*:*:*:*:*',
    'cpe:2.3:o:apple:macos:14.2:*:*:*:*:*:*:*',
    'cpe:2.3:o:apple:macos:13.6:*:*:*:*:*:*:*',
    'cpe:2.3:o:apple:ios:17.2:*:*:*:*:*:*:*',
    'cpe:2.3:o:apple:ipados:17.2:*:*:*:*:*:*:*',
    'cpe:2.3:o:canonical:ubuntu_linux:22.04:*:*:*:lts:*:*:*',
    'cpe:2.3:o:canonical:ubuntu_linux:20.04:*:*:*:lts:*:*:*',
    'cpe:2.3:o:debian:debian_linux:12:*:*:*:*:*:*:*',
    'cpe:2.3:o:debian:debian_linux:11:*:*:*:*:*:*:*',
    'cpe:2.3:o:redhat:enterprise_linux:9.0:*:*:*:*:*:*:*',
    'cpe:2.3:o:redhat:enterprise_linux:8.0:*:*:*:*:*:*:*',
    'cpe:2.3:o:centos:centos:8:*:*:*:*:*:*:*',
    'cpe:2.3:o:centos:centos:7:*:*:*:*:*:*:*',
    'cpe:2.3:o:fedoraproject:fedora:39:*:*:*:*:*:*:*',
    'cpe:2.3:o:fedoraproject:fedora:38:*:*:*:*:*:*:*',
    'cpe:2.3:o:opensuse:leap:15.5:*:*:*:*:*:*:*',
    'cpe:2.3:o:oracle:linux:8:*:*:*:*:*:*:*',
    'cpe:2.3:o:google:android:14.0:*:*:*:*:*:*:*',
    'cpe:2.3:o:google:android:13.0:*:*:*:*:*:*:*',
    'cpe:2.3:o:freebsd:freebsd:14.0:*:*:*:*:*:*:*',
    'cpe:2.3:o:netbsd:netbsd:10.0:*:*:*:*:*:*:*',
    
    # Hardware (h)
    'cpe:2.3:h:cisco:catalyst_9300:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:cisco:catalyst_2960:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:cisco:nexus_9000:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:hp:laserjet_pro_m404:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:hp:officejet_pro_9015:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:dell:poweredge_r750:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:dell:poweredge_r640:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:lenovo:thinkpad_x1_carbon:gen11:*:*:*:*:*:*:*',
    'cpe:2.3:h:lenovo:thinkpad_t14:gen3:*:*:*:*:*:*:*',
    'cpe:2.3:h:apple:macbook_pro:2023:*:*:*:*:*:*:*',
    'cpe:2.3:h:apple:macbook_air:2023:*:*:*:*:*:*:*',
    'cpe:2.3:h:apple:iphone_15:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:apple:ipad_pro:2023:*:*:*:*:*:*:*',
    'cpe:2.3:h:samsung:galaxy_s23:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:samsung:galaxy_tab_s9:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:intel:core_i9-13900k:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:intel:core_i7-13700k:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:amd:ryzen_9_7950x:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:amd:ryzen_7_7700x:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:nvidia:geforce_rtx_4090:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:nvidia:geforce_rtx_4080:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:asus:rog_strix_b650e:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:juniper:srx4100:*:*:*:*:*:*:*:*',
    'cpe:2.3:h:fortinet:fortigate_600e:*:*:*:*:*:*:*:*'
]

def parse_cpe_uri(cpe_string):
    """
    Parse CPE URI format (cpe:2.3:a:vendor:product:version:...)
    Returns dict with parsed components
    """
    try:
        if not cpe_string.startswith('cpe:'):
            return None
        
        parts = cpe_string.split(':')
        if len(parts) < 5:
            return None
        
        # Extract category (part type: a, o, h)
        part_type = parts[2] if len(parts) > 2 else ''
        category_map = {
            'a': 'Application',
            'o': 'Operating System',
            'h': 'Hardware'
        }
        category = category_map.get(part_type, 'Unknown')
        
        result = {
            'cpe': cpe_string,
            'category': category,
            'category_code': part_type,
            'vendor': parts[3] if len(parts) > 3 and parts[3] != '*' else '',
            'product': parts[4] if len(parts) > 4 and parts[4] != '*' else '',
            'version': parts[5] if len(parts) > 5 and parts[5] != '*' else '',
            'update': parts[6] if len(parts) > 6 and parts[6] != '*' else '',
            'edition': parts[7] if len(parts) > 7 and parts[7] != '*' else '',
            'language': parts[8] if len(parts) > 8 and parts[8] != '*' else '',
            'sw_edition': parts[9] if len(parts) > 9 and parts[9] != '*' else '',
            'target_sw': parts[10] if len(parts) > 10 and parts[10] != '*' else '',
            'target_hw': parts[11] if len(parts) > 11 and parts[11] != '*' else '',
        }
        
        # Create other fields description
        other_fields = []
        if result['update']:
            other_fields.append(f"Update: {result['update']}")
        if result['edition']:
            other_fields.append(f"Edition: {result['edition']}")
        if result['language']:
            other_fields.append(f"Language: {result['language']}")
        if result['sw_edition']:
            other_fields.append(f"SW Edition: {result['sw_edition']}")
        if result['target_sw']:
            other_fields.append(f"Target SW: {result['target_sw']}")
        if result['target_hw']:
            other_fields.append(f"Target HW: {result['target_hw']}")
        
        result['other_fields'] = ', '.join(other_fields) if other_fields else 'None'
        
        return result
    except Exception as e:
        print(f"Error parsing CPE: {e}")
        return None

def generate_installation_metadata():
    """Generate simulated installation metadata"""
    # Random size between 10MB and 2000MB
    size_mb = round(random.uniform(10, 2000), 2)
    
    # Random installation date within last 2 years
    days_ago = random.randint(0, 730)
    install_date = (datetime.now() - timedelta(days=days_ago)).strftime('%Y-%m-%d')
    
    # Installation location (default C drive)
    locations = [
        'C:\\Program Files\\',
        'C:\\Program Files (x86)\\',
        'C:\\Users\\Public\\',
        'C:\\ProgramData\\'
    ]
    location = random.choice(locations)
    
    return {
        'size_mb': size_mb,
        'install_date': install_date,
        'install_location': location
    }

def validate_cpe_with_nvd(cpe_string):
    """
    Validate CPE against NVD CPE dictionary
    Returns True if valid, False otherwise
    """
    try:
        # Try to search for the CPE in NVD
        # Note: In production, you would use NVD API with proper rate limiting
        # For this implementation, we'll do basic validation
        
        # Basic format validation
        if not cpe_string.startswith('cpe:2.3:'):
            return False
        
        parts = cpe_string.split(':')
        if len(parts) < 6:
            return False
        
        # Check if vendor and product are not empty
        if parts[3] == '*' or parts[4] == '*':
            return False
        
        return True
    except Exception as e:
        print(f"Error validating CPE: {e}")
        return False

def search_nvd_cpe(vendor, product, version=''):
    """
    Search NVD for CPE entries matching vendor, product, and optionally version
    Returns list of matching CPEs
    """
    try:
        # Note: This is a simplified implementation
        # In production, use the official NVD CPE API: https://nvd.nist.gov/developers/products
        
        # For demonstration, generate some sample CPEs
        cpes = []
        
        # Generate a few variations
        base_cpe = f"cpe:2.3:a:{vendor}:{product}:{version if version else '*'}"
        for i in range(3):
            v = version if version else f"{random.randint(1,10)}.{random.randint(0,9)}.{random.randint(0,9)}"
            cpe = f"cpe:2.3:a:{vendor}:{product}:{v}:*:*:*:*:*:*:*"
            cpes.append(cpe)
        
        return cpes
    except Exception as e:
        print(f"Error searching NVD: {e}")
        return []

def generate_random_cpe():
    """Generate random CPE with metadata"""
    vendor = random.choice(COMMON_VENDORS)
    
    # Generate product name
    if random.random() > 0.5:
        product = f"{random.choice(PRODUCT_PREFIXES)}_{random.choice(PRODUCT_TYPES)}"
    else:
        product = f"{vendor}_{random.choice(PRODUCT_TYPES)}"
    
    # Generate version
    major = random.randint(1, 20)
    minor = random.randint(0, 9)
    patch = random.randint(0, 99)
    version = f"{major}.{minor}.{patch}"
    
    # Create CPE string
    cpe_string = f"cpe:2.3:a:{vendor}:{product}:{version}:*:*:*:*:*:*:*"
    
    # Check if similar CPE exists and adjust if needed
    similar_cpes = search_nvd_cpe(vendor, product)
    if similar_cpes and random.random() > 0.5:
        # Sometimes use a similar existing CPE
        cpe_string = random.choice(similar_cpes)
    
    return {
        'vendor': vendor,
        'product': product,
        'version': version,
        'cpe': cpe_string,
        **generate_installation_metadata()
    }

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/api/auto-fetch-cpe', methods=['POST'])
def auto_fetch_cpe():
    """
    Auto-fetch CPE entries from CPE dictionary
    Expected input: count (number of CPE entries to fetch, default 10)
    Tries to evenly distribute h, o, a types
    """
    try:
        data = request.json
        count = data.get('count', 10)
        count = min(max(1, count), 100)  # Limit between 1 and 100
        
        # Check if data should be saved to database
        save_to_db = data.get('save_to_db', False)
        
        # Separate CPEs by category
        cpe_by_category = {'a': [], 'o': [], 'h': []}
        for cpe_string in CPE_DICTIONARY:
            parts = cpe_string.split(':')
            if len(parts) > 2:
                category = parts[2]
                if category in cpe_by_category:
                    cpe_by_category[category].append(cpe_string)
        
        # Calculate distribution for even split
        per_category = count // 3
        remainder = count % 3
        
        selected_cpes = []
        
        # Select evenly from each category
        for idx, (category, cpes) in enumerate(cpe_by_category.items()):
            # Add extra items to first categories for remainder
            category_count = per_category + (1 if idx < remainder else 0)
            if cpes:
                # Select random CPEs from this category
                available = min(category_count, len(cpes))
                selected = random.sample(cpes, available)
                selected_cpes.extend(selected)
        
        # If we don't have enough CPEs, fill from any available
        if len(selected_cpes) < count:
            all_cpes = [cpe for cpes in cpe_by_category.values() for cpe in cpes]
            remaining_cpes = [cpe for cpe in all_cpes if cpe not in selected_cpes]
            if remaining_cpes:
                additional = min(count - len(selected_cpes), len(remaining_cpes))
                selected_cpes.extend(random.sample(remaining_cpes, additional))
        
        # Shuffle to randomize order
        random.shuffle(selected_cpes)
        
        results = []
        for cpe_string in selected_cpes:
            # Validate CPE
            if validate_cpe_with_nvd(cpe_string):
                # Parse CPE
                parsed = parse_cpe_uri(cpe_string)
                if parsed:
                    # Add installation metadata
                    metadata = generate_installation_metadata()
                    results.append({**parsed, **metadata})
        
        # Save to database if requested
        if save_to_db and results:
            db_result = save_multiple_cpe_to_database(results)
            return jsonify({
                'data': results,
                'database': {
                    'saved': True,
                    'success_count': db_result['success'],
                    'failed_count': db_result['failed']
                }
            })
        
        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/fetch-cpe', methods=['POST'])
def fetch_cpe():
    """
    Fetch and parse CPE from user input
    Expected input: CPE string or search parameters
    """
    try:
        data = request.json
        cpe_input = data.get('cpe_string', '')
        
        if not cpe_input:
            return jsonify({'error': 'CPE string is required'}), 400
        
        # Validate CPE
        if not validate_cpe_with_nvd(cpe_input):
            return jsonify({'error': 'Invalid CPE format or CPE not found in dictionary'}), 400
        
        # Parse CPE
        parsed = parse_cpe_uri(cpe_input)
        if not parsed:
            return jsonify({'error': 'Failed to parse CPE'}), 400
        
        # Add installation metadata
        metadata = generate_installation_metadata()
        result = {**parsed, **metadata}
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/search-cpe', methods=['POST'])
def search_cpe():
    """
    Search for CPE entries based on vendor/product
    """
    try:
        data = request.json
        vendor = data.get('vendor', '')
        product = data.get('product', '')
        
        if not vendor or not product:
            return jsonify({'error': 'Vendor and product are required'}), 400
        
        # Search for CPEs
        cpes = search_nvd_cpe(vendor, product)
        
        results = []
        for cpe in cpes[:10]:  # Limit to 10 results
            parsed = parse_cpe_uri(cpe)
            if parsed:
                metadata = generate_installation_metadata()
                results.append({**parsed, **metadata})
        
        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate-random', methods=['POST'])
def generate_random():
    """
    Generate random CPE entries
    """
    try:
        data = request.json
        count = data.get('count', 5)
        count = min(count, 50)  # Limit to 50 entries
        
        results = []
        for _ in range(count):
            cpe_data = generate_random_cpe()
            # Parse the generated CPE to get all fields including category
            parsed = parse_cpe_uri(cpe_data['cpe'])
            if parsed:
                # Merge parsed data with generated metadata
                results.append({**parsed, **generate_installation_metadata()})
            else:
                # Fallback to original data if parsing fails
                results.append(cpe_data)
        
        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-csv', methods=['POST'])
def export_csv():
    """
    Export CPE data to CSV
    """
    try:
        data = request.json
        cpe_data = data.get('data', [])
        
        if not cpe_data:
            return jsonify({'error': 'No data to export'}), 400
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'CPE', 'Category', 'Vendor', 'Product', 'Version', 
            'Other Fields', 'Size (MB)', 'Install Date', 'Install Location'
        ])
        
        # Write data
        for item in cpe_data:
            writer.writerow([
                item.get('cpe', ''),
                item.get('category', ''),
                item.get('vendor', ''),
                item.get('product', ''),
                item.get('version', ''),
                item.get('other_fields', ''),
                item.get('size_mb', ''),
                item.get('install_date', ''),
                item.get('install_location', '')
            ])
        
        # Prepare response
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'cpe_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-xlsx', methods=['POST'])
def export_xlsx():
    """
    Export CPE data to XLSX (Excel)
    """
    try:
        data = request.json
        cpe_data = data.get('data', [])
        
        if not cpe_data:
            return jsonify({'error': 'No data to export'}), 400
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "CPE Data"
        
        # Define headers
        headers = [
            'CPE', 'Category', 'Vendor', 'Product', 'Version', 
            'Other Fields', 'Size (MB)', 'Install Date', 'Install Location'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, item in enumerate(cpe_data, 2):
            ws.cell(row=row_num, column=1, value=item.get('cpe', ''))
            ws.cell(row=row_num, column=2, value=item.get('category', ''))
            ws.cell(row=row_num, column=3, value=item.get('vendor', ''))
            ws.cell(row=row_num, column=4, value=item.get('product', ''))
            ws.cell(row=row_num, column=5, value=item.get('version', ''))
            ws.cell(row=row_num, column=6, value=item.get('other_fields', ''))
            ws.cell(row=row_num, column=7, value=item.get('size_mb', ''))
            ws.cell(row=row_num, column=8, value=item.get('install_date', ''))
            ws.cell(row=row_num, column=9, value=item.get('install_location', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'cpe_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-json', methods=['POST'])
def export_json():
    """
    Export CPE data to JSON
    """
    try:
        data = request.json
        cpe_data = data.get('data', [])
        
        if not cpe_data:
            return jsonify({'error': 'No data to export'}), 400
        
        # Format the data for JSON export
        json_output = {
            'export_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_entries': len(cpe_data),
            'data': cpe_data
        }
        
        # Convert to JSON string with proper formatting
        json_string = json.dumps(json_output, ensure_ascii=False, indent=2)
        
        return send_file(
            io.BytesIO(json_string.encode('utf-8')),
            mimetype='application/json',
            as_attachment=True,
            download_name=f'cpe_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-connections', methods=['GET'])
def get_db_connections():
    """取得所有已儲存的資料庫連線"""
    try:
        connections = load_db_connections()
        # 隱藏密碼資訊
        safe_connections = {}
        for name, config in connections.items():
            safe_config = config.copy()
            if 'password' in safe_config:
                safe_config['password'] = '***' if safe_config['password'] else ''
            safe_connections[name] = safe_config
        
        current_config = get_current_db_config()
        return jsonify({
            'connections': safe_connections,
            'current_connection': current_config.get('name', 'default') if current_config else 'default'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-connections', methods=['POST'])
def add_db_connection():
    """新增或更新資料庫連線設定"""
    try:
        data = request.json
        name = data.get('name')
        
        if not name:
            return jsonify({'error': '連線名稱為必填項目'}), 400
        
        # 驗證必要欄位
        required_fields = ['server', 'database']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} 為必填項目'}), 400
        
        # 載入現有連線
        connections = load_db_connections()
        
        # 建立連線配置
        connection_config = {
            'name': name,
            'server': data.get('server'),
            'database': data.get('database'),
            'username': data.get('username', ''),
            'password': data.get('password', ''),
            'trusted_connection': data.get('trusted_connection', False)
        }
        
        # 儲存連線
        connections[name] = connection_config
        
        if save_db_connections(connections):
            return jsonify({
                'success': True,
                'message': f'連線 "{name}" 已儲存'
            })
        else:
            return jsonify({'error': '儲存連線失敗'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-connections/<name>', methods=['PUT'])
def update_db_connection(name):
    """更新資料庫連線設定"""
    try:
        data = request.json
        connections = load_db_connections()
        
        if name not in connections:
            return jsonify({'error': f'連線 "{name}" 不存在'}), 404
        
        # 更新連線配置
        connection_config = connections[name]
        if 'server' in data:
            connection_config['server'] = data['server']
        if 'database' in data:
            connection_config['database'] = data['database']
        if 'username' in data:
            connection_config['username'] = data['username']
        if 'password' in data:
            connection_config['password'] = data['password']
        if 'trusted_connection' in data:
            connection_config['trusted_connection'] = data['trusted_connection']
        
        connections[name] = connection_config
        
        if save_db_connections(connections):
            return jsonify({
                'success': True,
                'message': f'連線 "{name}" 已更新'
            })
        else:
            return jsonify({'error': '更新連線失敗'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-connections/<name>', methods=['DELETE'])
def delete_db_connection(name):
    """刪除資料庫連線設定"""
    try:
        connections = load_db_connections()
        
        if name not in connections:
            return jsonify({'error': f'連線 "{name}" 不存在'}), 404
        
        del connections[name]
        
        if save_db_connections(connections):
            return jsonify({
                'success': True,
                'message': f'連線 "{name}" 已刪除'
            })
        else:
            return jsonify({'error': '刪除連線失敗'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-connections/test', methods=['POST'])
def test_connection():
    """測試資料庫連線"""
    try:
        data = request.json
        
        # 驗證必要欄位
        required_fields = ['server', 'database']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} 為必填項目'}), 400
        
        # 建立測試配置
        test_config = {
            'server': data.get('server'),
            'database': data.get('database'),
            'username': data.get('username', ''),
            'password': data.get('password', ''),
            'trusted_connection': data.get('trusted_connection', False)
        }
        
        success, message = test_db_connection(test_config)
        
        return jsonify({
            'success': success,
            'message': message
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db-connections/set-current', methods=['POST'])
def set_current_connection():
    """設定目前使用的資料庫連線"""
    try:
        data = request.json
        name = data.get('name')
        
        if not name:
            return jsonify({'error': '連線名稱為必填項目'}), 400
        
        connections = load_db_connections()
        
        if name not in connections:
            return jsonify({'error': f'連線 "{name}" 不存在'}), 404
        
        set_current_db_config(connections[name])
        
        return jsonify({
            'success': True,
            'message': f'已切換至連線 "{name}"'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Note: For production deployment, set debug=False and use a production WSGI server
    import os
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=5000)
